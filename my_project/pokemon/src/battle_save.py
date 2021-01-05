import openpyxl, logging, os
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(pathname)s - %(message)s', filename='log/battle_record.log')      # 初期設定 -> logging.debug() でログのターミナル表示ができる
                    # 変数「level」以上のログを表示, ログの表示形式の作成（ログの発生時間 - ログのレベル - ログの内容（メッセージ））, ログファイルの出力(ターミナルには出力されなくなる)
# logging.disable(logging.CRITICAL)       # ログの無効化（以下のログを表示しなくなる）
# 重要度によって使い分ける debug() -> info() -> warning() -> error() -> critical()
# LogRecord属性(formatの記述について)：https://docs.python.org/ja/3/library/logging.html#logrecord-attributes

project_info = """
\n---プロジェクト概要---
・
\n---条件---
・
\n---使い方--- 
・
-----------------------\n
"""

def battle_save():
  # ファイルを開く
  workbook = openpyxl.load_workbook( '/Users/nakamurakoyo/Desktop/pokemon_battle_record.xlsx' )
  logging.debug('\'pokemon_battle_record.xlsx\' sheet is {}'.format(workbook.get_sheet_names()))
  print('--------\nOpened File\n----------\n')

  # 読み込みデータの入力確認(勝ち負けで判定)
  read_sheet = workbook.get_sheet_by_name('input_template')
  print('勝ち負け：{}'.format(read_sheet['E7'].value))
  if read_sheet['E7'].value == None:
    print('勝ち負けが分かりません')
    return
  # 書き込み場所の始点確認
  write_sheet = workbook.get_sheet_by_name('S14')
  n_start = 2
  while write_sheet.cell(column=1, row=n_start).value != None:
    n_start += 7
  # データコピー
  for i in range(6):
    for j in range(5):
      write_sheet.cell(column=j+1, row=n_start+i).value = read_sheet.cell(column=j+1, row=i+2).value
  print('--------\nCopied\n----------\n')

  workbook.save('/Users/nakamurakoyo/Desktop/pokemon_battle_record.xlsx')
  ################# セーブしたら開きたいな #####################
  print('--------\nSaved\n----------\n')

  

# メイン
if __name__ == "__main__":
  print('\n--------\nStart\n----------\n')
  battle_save()
  print('--------\nFinish\n----------\n')